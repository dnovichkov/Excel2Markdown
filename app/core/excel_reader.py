"""Excel file reading module with support for .xls and .xlsx formats."""

from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO, List, TypedDict, Union

from loguru import logger

from app.core.exceptions import (
    EmptyFileError,
    EmptySheetError,
    InvalidFileFormatError,
)


class SheetData(TypedDict):
    """Type definition for sheet data structure."""

    sheetname: str
    headers: List[str]
    data: List[List[Any]]


def detect_excel_format(filename: str) -> str:
    """
    Detect Excel format based on file extension.

    Args:
        filename: Name of the file with extension.

    Returns:
        Format string: 'xls' or 'xlsx'.

    Raises:
        InvalidFileFormatError: If the file extension is not supported.
    """
    ext = Path(filename).suffix.lower()
    if ext == ".xls":
        return "xls"
    elif ext == ".xlsx":
        return "xlsx"
    else:
        raise InvalidFileFormatError(
            f"Unsupported file format: {ext}. Supported formats: .xls, .xlsx"
        )


def read_excel_xls(
    file_content: Union[bytes, BinaryIO],
    use_headers: bool = True,
) -> List[SheetData]:
    """
    Read data from .xls file using xlrd.

    Args:
        file_content: File content as bytes or file-like object.
        use_headers: If True, first row is treated as headers.

    Returns:
        List of SheetData dictionaries with sheet data.

    Raises:
        EmptyFileError: If the file contains no sheets.
        InvalidFileFormatError: If the file cannot be read.
    """
    import xlrd

    try:
        if isinstance(file_content, bytes):
            workbook = xlrd.open_workbook(file_contents=file_content)
        else:
            content = file_content.read()
            workbook = xlrd.open_workbook(file_contents=content)
    except xlrd.biffh.XLRDError as e:
        raise InvalidFileFormatError(f"Cannot read .xls file: {e}")

    sheet_names = workbook.sheet_names()
    if not sheet_names:
        raise EmptyFileError("Excel file contains no sheets")

    result: List[SheetData] = []

    for sheet_name in sheet_names:
        sheet = workbook.sheet_by_name(sheet_name)
        num_cols = sheet.ncols
        num_rows = sheet.nrows

        if num_rows == 0:
            logger.warning("Empty sheet skipped: {}", sheet_name)
            continue

        headers: List[str] = []
        data: List[List[Any]] = []
        start_row_idx = 0

        if use_headers and num_rows > 0:
            start_row_idx = 1
            for col_idx in range(num_cols):
                cell_value = sheet.cell(0, col_idx).value
                headers.append(str(cell_value) if cell_value else "")

        for row_idx in range(start_row_idx, num_rows):
            row_data: List[Any] = []
            for col_idx in range(num_cols):
                cell_value = sheet.cell(row_idx, col_idx).value
                row_data.append(cell_value)
            data.append(row_data)

        result.append(
            SheetData(
                sheetname=sheet_name,
                headers=headers,
                data=data,
            )
        )

    return result


def read_excel_xlsx(
    file_content: Union[bytes, BinaryIO],
    use_headers: bool = True,
) -> List[SheetData]:
    """
    Read data from .xlsx file using openpyxl.

    Args:
        file_content: File content as bytes or file-like object.
        use_headers: If True, first row is treated as headers.

    Returns:
        List of SheetData dictionaries with sheet data.

    Raises:
        EmptyFileError: If the file contains no sheets.
        InvalidFileFormatError: If the file cannot be read.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        if isinstance(file_content, bytes):
            file_obj = BytesIO(file_content)
        else:
            file_obj = file_content

        workbook = load_workbook(filename=file_obj, read_only=True, data_only=True)
    except InvalidFileException as e:
        raise InvalidFileFormatError(f"Cannot read .xlsx file: {e}")
    except Exception as e:
        raise InvalidFileFormatError(f"Error reading .xlsx file: {e}")

    sheet_names = workbook.sheetnames
    if not sheet_names:
        raise EmptyFileError("Excel file contains no sheets")

    result: List[SheetData] = []

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            logger.warning("Empty sheet skipped: {}", sheet_name)
            continue

        # Determine actual column count (exclude trailing None columns)
        max_cols = 0
        for row in rows:
            for i, cell in enumerate(row):
                if cell is not None:
                    max_cols = max(max_cols, i + 1)

        if max_cols == 0:
            logger.warning("Sheet with no data skipped: {}", sheet_name)
            continue

        headers: List[str] = []
        data: List[List[Any]] = []

        if use_headers and rows:
            header_row = rows[0][:max_cols]
            headers = [str(cell) if cell is not None else "" for cell in header_row]
            data_rows = rows[1:]
        else:
            data_rows = rows

        for row in data_rows:
            row_data = list(row[:max_cols])
            # Replace None with empty string for consistency
            row_data = [cell if cell is not None else "" for cell in row_data]
            data.append(row_data)

        result.append(
            SheetData(
                sheetname=sheet_name,
                headers=headers,
                data=data,
            )
        )

    workbook.close()
    return result


def get_excel_data(
    file_content: Union[bytes, BinaryIO],
    filename: str,
    use_headers: bool = True,
) -> List[SheetData]:
    """
    Read Excel file and extract data from all sheets.

    This is the main entry point for reading Excel files. It automatically
    detects the format and uses the appropriate reader.

    Args:
        file_content: File content as bytes or file-like object.
        filename: Original filename (used to detect format).
        use_headers: If True, first row is treated as headers.

    Returns:
        List of SheetData dictionaries with sheet data.

    Raises:
        InvalidFileFormatError: If format is not supported or file is invalid.
        EmptyFileError: If the file contains no data.
    """
    file_format = detect_excel_format(filename)

    logger.info("Reading Excel file: {} (format: {})", filename, file_format)

    if file_format == "xls":
        sheets = read_excel_xls(file_content, use_headers)
    else:
        sheets = read_excel_xlsx(file_content, use_headers)

    if not sheets:
        raise EmptyFileError("Excel file contains no data")

    logger.info("Successfully read {} sheet(s) from {}", len(sheets), filename)
    return sheets


def get_excel_data_from_path(
    file_path: str,
    use_headers: bool = True,
) -> List[SheetData]:
    """
    Read Excel file from filesystem path.

    Convenience function for reading files from disk.

    Args:
        file_path: Path to the Excel file.
        use_headers: If True, first row is treated as headers.

    Returns:
        List of SheetData dictionaries with sheet data.
    """
    path = Path(file_path)
    with open(path, "rb") as f:
        content = f.read()
    return get_excel_data(content, path.name, use_headers)
